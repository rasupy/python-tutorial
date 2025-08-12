import openpyxl as excel
import random

# 当たりのシートの番号を決める
atari = random.randint(1, 100)

# 新規ブックを作成
book = excel.Workbook()
book.active["B2"] = "当たりが書かれたシートを探そう"

# 繰り返し100回シートを作成する
for i in range(1, 101):
    # 新規シートを作成
    sname = str(i) + "番"
    sheet = book.create_sheet(title = sname)
    # シートに書き込む単語を決定
    word = "ハズレ"
    if i == atari: word = "当たり"
    # インパクトがあるようにwordで画面を埋める
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y + 1, x + 1)
            c.value = word

# ファイルに保存
book.save("game100.xlsx")
print("ok, atari=", atari)