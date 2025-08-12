# openpyxl を取り込む
import openpyxl as excel

# 新規ワークブックを作成
book = excel.Workbook()

# 既存のワークブックをファイルから開く
book = excel.load_workbook()

# ワークブックを開く（式があれば展開して開く）
book = excel.load_workbook(
    "ファイル名.xlsx", data_only=True)

# ワークをブックを明示的に閉じる
book.close()

# アクティブなシートを得る
sheet = book.active

# 任意の箇所にあるワークシートを得る（ 0 起点 ）
sheet = book.worksheets[1] # シート番号

# シート名（sheet1など）を指定して取得
sheet = book["シート名"]

# ブック内のシート名の一覧を得る
print( book.sheetnames )

# 新規シートを作成
sheet = book.create_sheet(title="シート名")

# 既存のシートをコピーして得る
sheet = book.copy_worksheet(book["シート名"])

# シート名を変更する
sheet.title = "新しい名前"

# シートを削除
book.remove(book["シート名"])