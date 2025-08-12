import openpyxl as excel

book = excel.load_workbook("test100.xlsx")
sheet = book.active

print(" --- for文を利用して1つずつ範囲内のセルを得る方法 --- ")
for y in range(2, 5):
    r = []
    for x in range(2, 5):
        v = sheet.cell(row = y, column = x).value
        r.append(v)
    print(r)

print(" --- ワークシートの特定の範囲を取り出す方法 ---")
for row in sheet["B2":"D4"]:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)

print(" --- 上の省略版 --- ")
for row in sheet["B2":"D4"]:
    print([c.value for c in row])

print(" --- iter_rows で繰り返し指定範囲を得る方法 --- ")
# イテレータを取得する
it = sheet.iter_rows(
    min_row = 2, min_col = 2,
    max_row = 4, max_col = 4
)
# for文で繰り返し値を得る
for row in it:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)